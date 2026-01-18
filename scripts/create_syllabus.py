"""
강의계획서(syllabus.xlsx) 생성 스크립트
contents.md 내용을 기반으로 Excel 형식의 강의계획서 생성
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_syllabus():
    """강의계획서 Excel 파일 생성"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "강의계획서"

    # 스타일 정의
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    normal_font = Font(size=11)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 과목 정보 섹션
    ws.merge_cells('A1:F1')
    ws['A1'] = '딥러닝 자연어처리 강의계획서'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    # 기본 정보
    info_data = [
        ('과목명', '딥러닝 자연어처리'),
        ('부제', '인문사회과학을 위한 AI 언어 모델의 이론과 응용: LLM 파인튜닝 중심'),
        ('학점', '3학점'),
        ('대상', '컴퓨터공학/AI 전공 학부생 (3~4학년)'),
        ('학기', '2025년 1학기'),
    ]

    row = 3
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = thin_border
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = value
        ws[f'B{row}'].border = thin_border
        row += 1

    row += 1  # 빈 줄

    # 주차별 강의 내용 헤더
    headers = ['주차', '주제', '세부 내용', '실습/과제', '비고']
    col_widths = [8, 35, 50, 40, 15]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    row += 1

    # 주차별 강의 데이터
    weekly_content = [
        {
            'week': '1',
            'topic': 'AI 시대의 개막과 개발 환경 준비',
            'content': '• 인공지능(AI)의 정의와 역사적 발전\n• AI, 머신러닝, 딥러닝의 관계와 차이점\n• 자연어처리(NLP)란 무엇인가\n• 언어 모델의 개념과 발전 과정\n• 주요 딥러닝 프레임워크 소개',
            'practice': '• Python 설치 및 가상환경 설정\n• PyTorch 설치 및 GPU 환경 확인\n• Google Colab 사용법\n\n[과제] 개인 개발 환경 구축 스크린샷 제출',
            'note': ''
        },
        {
            'week': '2',
            'topic': '언어 모델의 진화: 통계에서 신경망까지',
            'content': '• 언어 모델의 역할과 목적\n• 조건부 확률과 언어 모델\n• N-gram 모델의 원리\n• 희소성 문제와 스무딩 기법\n• Word Embedding (Word2Vec, GloVe, FastText)',
            'practice': '• N-gram 모델 직접 구현\n• 텍스트 데이터 전처리\n• Word2Vec 모델 활용\n\n[과제] Bigram 모델 구현 및 Perplexity 계산',
            'note': ''
        },
        {
            'week': '3',
            'topic': '딥러닝의 핵심: 신경망과 학습 원리',
            'content': '• 인공 신경망의 기본 구조\n• 퍼셉트론과 다층 퍼셉트론(MLP)\n• 활성화 함수 (ReLU, GELU, Softmax)\n• 손실 함수 (Cross-Entropy)\n• 경사 하강법과 역전파 알고리즘',
            'practice': '• PyTorch Tensor 기본 조작\n• Autograd 자동 미분\n• MLP 모델 직접 설계 및 학습\n\n[과제] MNIST 손글씨 분류 MLP 모델 구현',
            'note': ''
        },
        {
            'week': '4',
            'topic': 'PyTorch 기반 딥러닝 모델 개발 프로세스',
            'content': '• torch.nn.Module 활용\n• 데이터 처리 파이프라인 (Dataset, DataLoader)\n• 옵티마이저와 학습률 스케줄러\n• 모델 학습 루프 구현\n• 모델 평가 (Accuracy, F1-Score)',
            'practice': '• IMDb 영화 리뷰 데이터셋 전처리\n• MLP 기반 감성 분석 모델 구현\n\n[과제] 하이퍼파라미터 튜닝 성능 비교 분석',
            'note': ''
        },
        {
            'week': '5',
            'topic': '순차 데이터 처리: RNN과 LSTM/GRU',
            'content': '• 순차 데이터의 특성\n• RNN의 기본 구조와 작동 원리\n• 장기 의존성 문제와 기울기 소실\n• LSTM (Forget/Input/Output Gate)\n• GRU와 LSTM 비교\n• Sequence-to-Sequence 모델',
            'practice': '• PyTorch로 RNN/LSTM/GRU 구현\n• Character-level LSTM 언어 모델\n\n[과제] 셰익스피어 텍스트로 LSTM 텍스트 생성',
            'note': ''
        },
        {
            'week': '6',
            'topic': '혁신의 중심: Transformer 아키텍처',
            'content': '• Transformer 등장 배경\n• Attention 메커니즘 (Q, K, V)\n• Scaled Dot-Product Attention\n• Self-Attention과 Multi-Head Attention\n• Positional Encoding\n• Transformer Encoder/Decoder 구조',
            'practice': '• Self-Attention 단계별 구현\n• Attention Weights 시각화\n• Transformer Encoder 블록 구현\n\n[과제] Transformer Encoder로 텍스트 분류 모델 구현',
            'note': ''
        },
        {
            'week': '7',
            'topic': '중간고사',
            'content': '• AI 기초, 신경망, RNN/LSTM, Transformer 원리\n• PyTorch 기본, 모델 구현, 데이터 처리\n• Attention 메커니즘 설명, RNN vs Transformer 비교',
            'practice': '이론 평가 + 코딩 실습 + 서술형 문제',
            'note': '중간고사'
        },
        {
            'week': '8',
            'topic': '텍스트 속 숨겨진 주제 찾기: 임베딩 기반 토픽 모델링',
            'content': '• 토픽 모델링의 정의와 목적\n• LDA의 원리와 한계점\n• BERTopic 아키텍처 (5단계 파이프라인)\n• UMAP, HDBSCAN, c-TF-IDF\n• Dynamic/Guided/Hierarchical Topic Modeling',
            'practice': '• BERTopic 기본 사용법\n• 뉴스 기사 토픽 모델링\n• 토픽 시각화 및 트렌드 분석\n\n[프로젝트] 관심 분야 문서 토픽 모델링 분석 보고서',
            'note': ''
        },
        {
            'week': '9',
            'topic': 'LLM 시대 (1) - BERT 아키텍처와 활용',
            'content': '• Pre-training과 Fine-tuning 전략\n• BERT Encoder-only 구조\n• Masked Language Model (MLM)\n• Next Sentence Prediction (NSP)\n• WordPiece Tokenization\n• BERT 변형 모델 (RoBERTa, ALBERT, DistilBERT)\n• Hugging Face Transformers 라이브러리',
            'practice': '• Hugging Face Pipeline 사용법\n• BERT Tokenizer 활용\n• BERT 기반 텍스트 분류/NER\n• BERT 임베딩 추출 및 유사도 계산\n\n[프로젝트] 전공 분야 BERT 기반 분류 모델 구축',
            'note': ''
        },
        {
            'week': '10',
            'topic': 'LLM 시대 (2) - GPT 아키텍처와 생성 모델',
            'content': '• Autoregressive Language Modeling\n• GPT Decoder-only 구조\n• Causal Self-Attention\n• GPT-1, GPT-2, GPT-3 발전 과정\n• 텍스트 생성 (Greedy, Beam, Sampling)\n• Zero-shot, Few-shot, In-Context Learning\n• 프롬프트 엔지니어링 기초',
            'practice': '• GPT-2 모델 텍스트 생성\n• 디코딩 전략 비교 실험\n• Temperature, Top-k, Top-p 조정\n• Zero-shot, Few-shot 프롬프팅 실습\n\n[프로젝트] 도메인 특화 텍스트 생성기 프로토타입',
            'note': ''
        },
        {
            'week': '11',
            'topic': 'LLM 파인튜닝 (1) - 전이 학습과 Full Fine-tuning',
            'content': '• 전이 학습의 개념과 필요성\n• Feature Extraction vs Fine-tuning\n• 파인튜닝 태스크 (Classification, NER, QA)\n• 데이터셋 준비 및 불균형 처리\n• Hugging Face Trainer API\n• 하이퍼파라미터 튜닝 (LR, Batch Size, Warmup)\n• 과적합 방지 기법',
            'practice': '• 금융 뉴스 분류 BERT 파인튜닝\n• Trainer API 활용 모델 학습\n• 학습 과정 시각화 및 분석\n• 파인튜닝 전후 성능 비교\n\n[프로젝트] 전공 분야 특화 모델 파인튜닝 (1단계)',
            'note': ''
        },
        {
            'week': '12',
            'topic': 'LLM 파인튜닝 (2) - PEFT와 효율적 튜닝',
            'content': '• Full Fine-tuning의 한계\n• Parameter-Efficient Fine-Tuning (PEFT) 개요\n• LoRA (Low-Rank Adaptation) 심화\n• LoRA 하이퍼파라미터 (r, α, Target Modules)\n• QLoRA (Quantized LoRA)\n• 기타 PEFT 기법 (Prefix Tuning, Adapter)\n• Hugging Face PEFT 라이브러리',
            'practice': '• LoRA Config 설정 및 적용\n• Full Fine-tuning vs LoRA 성능 비교\n• Rank와 Alpha 값 실험\n• QLoRA 대형 모델 튜닝\n\n[프로젝트] LoRA 파인튜닝 및 성능 비교 보고서',
            'note': ''
        },
        {
            'week': '13',
            'topic': 'LLM 고급 응용 - RAG와 프롬프트 엔지니어링',
            'content': '• LLM의 한계 (Hallucination, 지식 한계)\n• 검색 증강 생성 (RAG) 개요\n• RAG 시스템 구성 요소 (Retriever, Generator)\n• Vector Database (FAISS, ChromaDB)\n• RAG 파이프라인 (청킹, 임베딩, 검색, 생성)\n• LangChain 프레임워크\n• 프롬프트 엔지니어링 심화 (CoT, Few-shot)',
            'practice': '• LangChain 기본 사용법\n• FAISS Vector Database 구축\n• 간단한 RAG 시스템 구현\n• OpenAI/Claude API 활용\n\n[프로젝트] RAG 기반 Q&A 시스템 구축 (선택)',
            'note': ''
        },
        {
            'week': '14',
            'topic': '최종 프로젝트 개발 및 발표 준비',
            'content': '• 프로젝트 발표 가이드라인\n• 모델 평가 및 보고서 작성\n• Error Analysis, Ablation Study\n• 시각화 및 결과 해석\n• 모델 배포 및 윤리적 고려사항',
            'practice': '• 개인/팀별 프로젝트 개발\n• 중간 점검 및 피드백 세션\n• 발표 자료 준비',
            'note': '프로젝트 개발'
        },
        {
            'week': '15',
            'topic': '기말고사 및 프로젝트 최종 발표',
            'content': '• LLM 아키텍처 (BERT, GPT)\n• 파인튜닝 기법 (Full vs PEFT)\n• LoRA 원리 및 하이퍼파라미터\n• RAG 시스템 구성\n• 프롬프트 엔지니어링',
            'practice': '• 이론 평가 (전반부)\n• 프로젝트 최종 발표 (후반부, 10-15분)',
            'note': '기말고사\n프로젝트 발표'
        },
    ]

    for item in weekly_content:
        ws.cell(row=row, column=1, value=item['week']).alignment = center_align
        ws.cell(row=row, column=2, value=item['topic']).alignment = left_align
        ws.cell(row=row, column=3, value=item['content']).alignment = left_align
        ws.cell(row=row, column=4, value=item['practice']).alignment = left_align
        ws.cell(row=row, column=5, value=item['note']).alignment = center_align

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = normal_font

        # 행 높이 조정
        ws.row_dimensions[row].height = 100
        row += 1

    row += 1

    # 평가 방식 섹션
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = '평가 방식'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font_white
    ws[f'A{row}'].alignment = center_align
    row += 1

    evaluation = [
        ('과제', '20%', '주차별 실습 과제 (8-10회)'),
        ('중간고사', '30%', '이론 및 코딩 시험'),
        ('기말고사', '20%', '종합 이론 평가'),
        ('최종 프로젝트', '30%', '개인/팀 LLM 응용 프로젝트'),
    ]

    for item, percent, desc in evaluation:
        ws.cell(row=row, column=1, value=item).border = thin_border
        ws.cell(row=row, column=2, value=percent).border = thin_border
        ws.merge_cells(f'C{row}:E{row}')
        ws.cell(row=row, column=3, value=desc).border = thin_border
        row += 1

    row += 1

    # 참고 자료 섹션
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = '참고 자료'
    ws[f'A{row}'].font = header_font_white
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = center_align
    row += 1

    references = [
        '• Hugging Face Documentation (https://huggingface.co/docs)',
        '• PyTorch Documentation (https://pytorch.org/docs)',
        '• BERTopic Documentation (https://maartengr.github.io/BERTopic/)',
        '• PEFT Documentation (https://huggingface.co/docs/peft)',
        '• "Natural Language Processing with Transformers" (Lewis Tunstall et al.)',
        '• "Dive into Deep Learning" (https://d2l.ai/)',
        '• Stanford CS224N: NLP with Deep Learning',
    ]

    for ref in references:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = ref
        ws[f'A{row}'].border = thin_border
        row += 1

    # 열 너비 설정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 15

    # 파일 저장
    output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'syllabus.xlsx')
    wb.save(output_path)
    print(f"강의계획서가 생성되었습니다: {output_path}")
    return output_path

if __name__ == "__main__":
    create_syllabus()
