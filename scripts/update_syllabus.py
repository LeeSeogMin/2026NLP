"""
syllabus.xlsx 내용 업데이트 스크립트
원본 포맷(3열: 주차, 학습목표, 교재명 및 페이지 정보)을 유지하면서 contents.md 내용으로 업데이트
"""

import openpyxl

def update_syllabus():
    """syllabus.xlsx 내용 업데이트"""

    wb = openpyxl.load_workbook('c:/Dev/2026NLP/syllabus.xlsx')
    ws = wb['Sheet1']

    # contents.md 기반 15주차 데이터
    weekly_data = [
        {
            'week': '01',
            'topic': 'AI 시대의 개막과 개발 환경 준비',
            'content': '- 인공지능(AI)의 정의와 역사적 발전\n- AI, 머신러닝, 딥러닝의 관계와 차이점\n- 자연어처리(NLP)와 언어 모델의 개념\n- 주요 딥러닝 프레임워크 및 Hugging Face 생태계 소개\n- 실습: Python/PyTorch 설치, Google Colab GPU 설정\n- 과제: 개인 개발 환경 구축 스크린샷 제출'
        },
        {
            'week': '02',
            'topic': '언어 모델의 진화: 통계에서 신경망까지',
            'content': '- 언어 모델의 역할과 조건부 확률\n- N-gram 모델의 원리와 한계 (희소성 문제, 스무딩)\n- 신경망 기반 언어 모델과 분산 표현\n- 단어 임베딩: Word2Vec, GloVe, FastText\n- 실습: N-gram 모델 구현, 텍스트 전처리, Word2Vec 활용\n- 과제: Bigram 모델 구현 및 Perplexity 계산'
        },
        {
            'week': '03',
            'topic': '딥러닝의 핵심: 신경망과 학습 원리',
            'content': '- 인공 신경망 기본 구조 (퍼셉트론, MLP)\n- 활성화 함수 (Sigmoid, ReLU, GELU, Softmax)\n- 손실 함수 (MSE, Cross-Entropy)\n- 경사 하강법과 역전파 알고리즘\n- 실습: PyTorch Tensor, Autograd, MLP 구현\n- 과제: MNIST 손글씨 분류 MLP 모델 구현'
        },
        {
            'week': '04',
            'topic': 'PyTorch 기반 딥러닝 모델 개발 프로세스',
            'content': '- torch.nn.Module 활용 모델 정의\n- 데이터 파이프라인 (Dataset, DataLoader)\n- 옵티마이저 (SGD, Adam)와 학습률 스케줄러\n- 학습/검증 루프 구현, 과적합 방지 기법\n- 실습: IMDb 감성 분석 MLP 모델 구현\n- 과제: 하이퍼파라미터 튜닝 성능 비교 분석'
        },
        {
            'week': '05',
            'topic': '순차 데이터 처리: RNN과 LSTM/GRU',
            'content': '- 순차 데이터 특성과 RNN 기본 구조\n- 장기 의존성 문제와 기울기 소실/폭주\n- LSTM (Forget/Input/Output Gate)\n- GRU와 Seq2Seq 모델\n- 실습: PyTorch RNN/LSTM/GRU 구현\n- 과제: Character-level LSTM 텍스트 생성'
        },
        {
            'week': '06',
            'topic': '혁신의 중심: Transformer 아키텍처',
            'content': '- Transformer 등장 배경 (RNN의 한계)\n- Attention 메커니즘 (Query, Key, Value)\n- Self-Attention과 Multi-Head Attention\n- Positional Encoding, Encoder/Decoder 구조\n- 실습: Self-Attention 구현, Attention 시각화\n- 과제: Transformer Encoder 텍스트 분류 모델'
        },
        {
            'week': '07',
            'topic': '중간고사',
            'content': '- 이론 평가: AI 기초, 신경망, RNN/LSTM, Transformer 원리\n- 코딩 실습: PyTorch 기본, 모델 구현, 데이터 처리\n- 서술형: Attention 메커니즘 설명, RNN vs Transformer 비교'
        },
        {
            'week': '08',
            'topic': '임베딩 기반 토픽 모델링: BERTopic',
            'content': '- 토픽 모델링 개요 (LDA vs BERTopic)\n- BERTopic 5단계 파이프라인 (Embedding→UMAP→HDBSCAN→c-TF-IDF)\n- Dynamic/Guided/Hierarchical Topic Modeling\n- 실습: BERTopic으로 뉴스 기사 토픽 모델링\n- 프로젝트: 관심 분야 문서 토픽 분석 보고서'
        },
        {
            'week': '09',
            'topic': 'LLM 시대 (1): BERT 아키텍처와 활용',
            'content': '- Pre-training/Fine-tuning 전략, Transfer Learning\n- BERT Encoder-only 구조 (MLM, NSP)\n- WordPiece Tokenization, BERT 변형 모델\n- Hugging Face Transformers 라이브러리\n- 실습: BERT 텍스트 분류, NER, 임베딩 추출\n- 프로젝트: 전공 분야 BERT 분류 모델 구축'
        },
        {
            'week': '10',
            'topic': 'LLM 시대 (2): GPT 아키텍처와 생성 모델',
            'content': '- Autoregressive LM, Decoder-only 구조\n- Causal Self-Attention, GPT 발전 과정\n- 텍스트 생성 (Greedy, Beam, Top-k, Top-p)\n- Zero-shot, Few-shot, In-Context Learning\n- 실습: GPT-2 텍스트 생성, 프롬프트 엔지니어링\n- 프로젝트: 도메인 특화 텍스트 생성기'
        },
        {
            'week': '11',
            'topic': 'LLM 파인튜닝 (1): 전이 학습과 Full Fine-tuning',
            'content': '- 전이 학습 개념, Feature Extraction vs Fine-tuning\n- 파인튜닝 태스크 (Classification, NER, QA)\n- 데이터셋 준비, Hugging Face Trainer API\n- 하이퍼파라미터 튜닝, 과적합 방지\n- 실습: BERT 금융 뉴스 분류 파인튜닝\n- 프로젝트: 전공 분야 특화 모델 파인튜닝 (1단계)'
        },
        {
            'week': '12',
            'topic': 'LLM 파인튜닝 (2): PEFT와 효율적 튜닝',
            'content': '- Full Fine-tuning 한계와 PEFT 개요\n- LoRA (Low-Rank Adaptation) 원리와 하이퍼파라미터\n- QLoRA (Quantized LoRA)\n- Hugging Face PEFT 라이브러리\n- 실습: LoRA vs Full Fine-tuning 성능 비교\n- 프로젝트: LoRA 파인튜닝 및 비교 보고서'
        },
        {
            'week': '13',
            'topic': 'LLM 고급 응용: RAG와 프롬프트 엔지니어링',
            'content': '- LLM 한계 (Hallucination, 지식 한계)\n- RAG 개요 (Retriever + Generator)\n- Vector Database (FAISS, ChromaDB)\n- LangChain 프레임워크, 프롬프트 엔지니어링 심화\n- 실습: RAG 시스템 구현, OpenAI/Claude API 활용\n- 프로젝트: RAG 기반 Q&A 시스템 구축 (선택)'
        },
        {
            'week': '14',
            'topic': '최종 프로젝트 개발 및 발표 준비',
            'content': '- 프로젝트 발표 가이드라인 및 평가 기준\n- 모델 평가, Error Analysis, Ablation Study\n- 시각화 및 결과 해석\n- 모델 배포 및 윤리적 고려사항\n- 실습: 프로젝트 개발, 피드백 세션, 발표 자료 준비'
        },
        {
            'week': '15',
            'topic': '기말고사 및 프로젝트 최종 발표',
            'content': '- 기말고사 (전반부): LLM 아키텍처, 파인튜닝, LoRA, RAG, 프롬프트 엔지니어링\n- 프로젝트 발표 (후반부, 10-15분): 문제 정의, 방법론, 실험 결과, 결론\n- 제출물: 보고서(PDF), 소스코드(GitHub), 발표자료(PPT)'
        },
    ]

    # 데이터 업데이트 (Row 2부터 Row 16까지)
    for i, data in enumerate(weekly_data):
        row = i + 2  # 헤더가 Row 1이므로 Row 2부터 시작
        ws.cell(row=row, column=1, value=data['week'])
        ws.cell(row=row, column=2, value=data['topic'])
        ws.cell(row=row, column=3, value=data['content'])

    # 파일 저장
    wb.save('c:/Dev/2026NLP/syllabus.xlsx')
    print("syllabus.xlsx 업데이트 완료")

if __name__ == "__main__":
    update_syllabus()
